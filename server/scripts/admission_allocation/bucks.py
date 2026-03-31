from __future__ import annotations

import io
import json
import re
import sys
from typing import Any

import requests

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore

from . import common

STATISTICS_PAGE_URL = (
    "https://www.buckinghamshire.gov.uk/schools-and-learning/schools-index/"
    "school-admissions/school-admissions-guides-policies-and-statistics/"
    "school-place-allocation-statistics/"
)

INGEST: list[dict[str, Any]] = [
    {
        "entry_year": 2026,
        "statistics_page_url": STATISTICS_PAGE_URL,
        "la_slug": "buckinghamshire",
        "rounds": [
            {
                "round_order": 1,
                "round_code": "march_national",
                "url": "https://www.buckinghamshire.gov.uk/documents/40617/Allocation_Profile_2026_-_FINAL_V2.xlsx",
            },
        ],
    },
    {
        "entry_year": 2025,
        "statistics_page_url": STATISTICS_PAGE_URL,
        "la_slug": "buckinghamshire",
        "rounds": [
            {
                "round_order": 1,
                "round_code": "march_national",
                "url": "https://www.buckinghamshire.gov.uk/documents/36157/Allocation_Profile_2025_-_FINAL.xlsx",
            },
            {
                "round_order": 2,
                "round_code": "realloc_april",
                "url": "https://www.buckinghamshire.gov.uk/documents/37366/Secondary_school_allocations_-_2_April_2025_Re-allocation_Round.xlsx",
            },
            {
                "round_order": 3,
                "round_code": "realloc_may",
                "url": "https://www.buckinghamshire.gov.uk/documents/37365/Secondary_school_allocations_-_21_May_2025_Second_Round.xlsx",
            },
        ],
    },
    {
        "entry_year": 2024,
        "statistics_page_url": STATISTICS_PAGE_URL,
        "la_slug": "buckinghamshire",
        "rounds": [
            {
                "round_order": 1,
                "round_code": "march_national",
                "url": "https://www.buckinghamshire.gov.uk/documents/33847/Allocation_Profile_2024_-_FINAL_rrgfYhf.xlsx",
            },
        ],
    },
    {
        "entry_year": 2023,
        "statistics_page_url": STATISTICS_PAGE_URL,
        "la_slug": "buckinghamshire",
        "rounds": [
            {
                "round_order": 1,
                "round_code": "march_national",
                "url": "https://www.buckinghamshire.gov.uk/documents/33846/ALLOCATION_PROFILE_2023__5ua2yRL.xlsx",
            },
        ],
    },
]


def split_profile_line_items(narrative: str) -> list[str]:
    text = (narrative or "").replace("\n", " ").strip()
    if not text:
        return []
    text = re.sub(r"\s+", " ", text)

    extras: list[str] = []
    pan_pat = re.compile(
        r"((?:\d+|[a-z]+)\s+place(?:s)?\s+allocated\s+over\s+PAN\b.*)$",
        re.IGNORECASE,
    )
    m = pan_pat.search(text)
    if m:
        extras.append(m.group(1).strip().rstrip("."))
        text = text[: m.start()].strip().rstrip(".").strip()

    if re.fullmatch(r"no\s+further\s+offers\s+made\.?", text, re.IGNORECASE):
        return ["No further offers made."]

    parts = re.split(r",\s*(?=[Rr]ule\s+\d+)", text)
    items = [p.strip().rstrip(".") for p in parts if p.strip()]
    for e in extras:
        if e and e not in items:
            items.append(e)
    return items


def download_xlsx(url: str) -> bytes:
    r = requests.get(url, timeout=120)
    r.raise_for_status()
    return r.content


def parse_grammar_rows(xlsx_bytes: bytes) -> tuple[str | None, str | None, list[tuple[str, str]]]:
    if not openpyxl:
        raise RuntimeError("openpyxl required for Bucks loader")
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active
    round_label = None
    r0 = ws.cell(1, 1).value
    r0s = str(r0).strip() if r0 is not None else ""
    row_start = 1
    if r0s and "GRAMMAR SCHOOLS" not in r0s.upper():
        round_label = r0s
        row_start = 2

    profile_heading = None
    rows_out: list[tuple[str, str]] = []
    in_grammar = False
    for r in range(row_start, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if a is None and b is None:
            continue
        a_str = str(a).replace("\n", " ").strip() if a is not None else ""
        b_str = str(b).strip() if b is not None else ""
        up = a_str.upper()
        if up.startswith("UPPER SCHOOLS"):
            break
        if up.startswith("GRAMMAR SCHOOLS") and b_str:
            in_grammar = True
            profile_heading = b_str
            continue
        if in_grammar and a_str and b_str:
            rows_out.append((a_str, b_str))
    return round_label, profile_heading, rows_out


def run(cur, *, dry_run: bool = False) -> tuple[int, list[str]]:
    if not openpyxl:
        print("Missing openpyxl: pip install openpyxl", file=sys.stderr)
        return 0, []

    pending: list[tuple] = []
    for cohort in INGEST:
        entry_year = cohort["entry_year"]
        la_slug = cohort["la_slug"]
        stats_url = cohort["statistics_page_url"]
        for rnd in cohort["rounds"]:
            data = download_xlsx(rnd["url"])
            round_label, profile_heading, grammar_rows = parse_grammar_rows(data)
            for xlsx_name, narrative in grammar_rows:
                line_items = split_profile_line_items(narrative)
                pending.append(
                    (
                        entry_year,
                        la_slug,
                        int(rnd["round_order"]),
                        str(rnd["round_code"]),
                        round_label or "",
                        profile_heading or "",
                        json.dumps(line_items),
                        rnd["url"],
                        stats_url,
                        xlsx_name.strip(),
                    )
                )

    if dry_run:
        print(f"[bucks] Parsed {len(pending)} school-round rows (preview):")
        for row in pending[:5]:
            li = json.loads(row[6]) if row[6] else []
            print(" ", row[-1], row[0], row[3], li[:2])
        return len(pending), []

    by_key = common.load_school_maps(cur)
    cur.execute("DELETE FROM admissions_allocation_history WHERE la_slug = %s", ("buckinghamshire",))
    n = 0
    misses: list[str] = []
    for (
        entry_year,
        la_slug,
        round_order,
        round_code,
        round_label,
        profile_heading,
        line_items_json,
        data_source_url,
        statistics_page_url,
        xlsx_name,
    ) in pending:
        resolved = common.resolve_school(xlsx_name, by_key)
        if not resolved:
            misses.append(xlsx_name)
            continue
        sid, _name, urn = resolved
        cur.execute(
            common.INSERT_SQL,
            (
                sid,
                entry_year,
                urn,
                la_slug,
                round_order,
                round_code,
                round_label,
                profile_heading,
                line_items_json,
                data_source_url,
                statistics_page_url,
            ),
        )
        n += 1
    common.touch_refresh_log(cur, "bucks_allocation_profiles_ingest")
    return n, misses
